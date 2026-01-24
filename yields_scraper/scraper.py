from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import Select
import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import numpy as np
import xlwings as xw

url = "https://gemelnet.cma.gov.il/views/dafmakdim.aspx"
download_path = "/Users/rotemjablonsky/analytics_work/yields_scraper"

def check_new_file(category_name, sub_category):
    existing_files = set(os.listdir(download_path))
    timeout = 30
    start_time = time.time()

    continue_loop = True
    while continue_loop:
        current_files = set(os.listdir(download_path))
        new_files = current_files - existing_files

        print(f'New file detected: {new_files}')

        for file in new_files:
            if file.endswith('.xls') or file.endswith('.xlsx'):
                print(f'New file found: {file}')
                new_name = f'{category_name}_{sub_category}.xls'
                old_path = os.path.join(download_path, file)
                new_path = os.path.join(download_path, new_name)
                os.rename(old_path, new_path)
                continue_loop = False
                break

        if time.time() - start_time > timeout:
            print('No new file found')
            driver.quit()
            break

        time.sleep(1)

def yields_extractor(kupot, category_name):
    for kupa in kupot:
        try:
            text_entry.send_keys(f'{kupa}')
            WebDriverWait(driver, 2).until(EC.text_to_be_present_in_element((By.ID, 'selActiveKupot'), f'{kupa}'))
            add_button = driver.find_element(By.ID, 'btnAdd')
            add_button.click()
            WebDriverWait(driver, 2).until(EC.text_to_be_present_in_element((By.ID, 'selSelectedKupot'), f'{kupa}'))
            text_entry.clear()
        except TimeoutException:
            print(f'waiting time for kupa: {kupa} over')
            text_entry.clear()
        except NoSuchElementException:
            print(f'kupa: {kupa} does not found')
            text_entry.clear()

    load_table_and_export(load_table=load_table, export_to_excel_button=export_to_excel_button, category_name=category_name, sub_category='yields')
    alpha_button.click()
    load_table_and_export(load_table=load_table, export_to_excel_button=export_to_excel_button, category_name=category_name, sub_category='alpha')
    yield_button.click()
    remove_kupot()

def load_table_and_export(load_table, export_to_excel_button, category_name, sub_category):
    load_table.click()
    time.sleep(8)
    export_to_excel_button.click()
    print('Export to excel button clicked, waiting for a new file...')
    check_new_file(category_name, sub_category)

def remove_kupot():
    for option in list(selected_dropdown.options):
        value = option.get_attribute("value")
        selected_dropdown.select_by_value(value)
        remove_button.click()
        time.sleep(0.2)

hishtalmut_clali = [
    'הראל השתלמות כללי',
    'כלל השתלמות כללי',
    'מגדל השתלמות כללי',
    'מנורה השתלמות כללי',
    'מיטב השתלמות כללי',
    'אנליסט השתלמות כללי',
    'הפניקס השתלמות כללי',
    'אלטשולר שחם השתלמות כללי',
    'ילין לפידות קרן השתלמות מסלול כללי',
    'מור השתלמות - כללי'
]

hishtalmut_stocks = [
    'מגדל השתלמות מניות',
    'מיטב השתלמות מניות',
    'אנליסט השתלמות מניות',
    'הפניקס השתלמות מניות',
    'מנורה השתלמות מניות',
    'כלל השתלמות מניות',
    'אלטשולר שחם השתלמות מניות',
    'הראל השתלמות מסלול מניות',
    'ילין לפידות קרן השתלמות מסלול מניות',
    'מור השתלמות - מניות'
]

hishtalmut_sp = [
    'מור השתלמות -עוקב מדד S&P 500',
    'מיטב השתלמות עוקב מדד S&P500',
    'אנליסט השתלמות עוקב מדד S&P500',#
    'מנורה השתלמות עוקב S&P500',#
    'מגדל השתלמות עוקב מדד S&P 500',#
    'אלטשולר שחם השתלמות עוקב מדד S&P 500',
    '13264', #הפניקס השתלמות עוקב מדד s&p 500
    '13342', #כלל השתלמות עוקב מדד s&p 500
    'הראל השתלמות - עוקב מדד s&p 500',#
    'ילין לפידות קרן השתלמות מסלול עוקב מדד s&p 500'#
]

gemel_clali = [
    'אנליסט קופת גמל להשקעה כללי',
    'הפניקס גמל להשקעה כללי',
    'מגדל גמל להשקעה כללי',
    'מיטב גמל להשקעה כללי',
    'הראל גמל להשקעה כללי',
    'מנורה מבטחים גמל להשקעה כללי',
    'ילין לפידות קופת גמל להשקעה מסלול כללי',
    'אלטשולר שחם חיסכון פלוס כללי',
    'כלל גמל לעתיד כללי',
    'מור גמל להשקעה- כללי'
]

gemel_stocks = [
    'אנליסט קופת גמל להשקעה מניות',
    'מיטב גמל להשקעה מניות',
    'מגדל גמל להשקעה מניות',
    'הפניקס גמל להשקעה מניות',
    'הראל גמל להשקעה מניות',
    'מנורה מבטחים גמל להשקעה מניות סחיר',
    'ילין לפידות קופת גמל להשקעה מסלול מניות',
    'אלטשולר שחם חיסכון פלוס מניות',
    'כלל גמל לעתיד מניות',
    'מור גמל להשקעה - מניות'
]

gemel_sp = [
    '13344', #כלל גמל לעתיד עוקב מדד s&p 500
    'הראל גמל להשקעה עוקב מדד s&p 500',#
    'מגדל גמל להשקעה עוקב מדד s&p500',#
    'ילין לפידות קופת גמל להשקעה מסלול עוקב מדד s&p 500',#
    'מור גמל להשקעה - עוקב מדד S&P500',#
    'הפניקס גמל להשקעה עוקב מדד S&P500',
    'מיטב גמל להשקעה עוקב מדד S&P 500',
    'אנליסט קופת גמל להשקעה עוקב מדד S&P500',#
    'מנורה מבטחים גמל להשקעה עוקב S&P500',#
    'אלטשולר שחם חיסכון פלוס עוקב מדד S&P 500'
]

options = Options()
options.add_experimental_option('prefs', {
    "download.default_directory": download_path,
    "download.prompt_for_download": False,
    "profile.default_content_settings.popups": 0,
    "directory_upgrade": True
})

driver = webdriver.Chrome(options=options)
driver.get(url)
time.sleep(2)

enter_button = driver.find_element(By.ID, 'knisa')
enter_button.click()
time.sleep(4)

load_table = driver.find_element(By.ID, 'cbDisplay')
export_to_excel_button = driver.find_element(By.ID, 'cbExcel')
yield_button = driver.find_element(By.ID, 'rbKupotDefault')
alpha_button = driver.find_element(By.ID, 'rbKupotAlpha')
text_entry = driver.find_element(By.ID, 'txtFilterKupot')
text_entry.clear()
selected_dropdown = Select(driver.find_element(By.ID, 'selSelectedKupot'))
remove_button = driver.find_element(By.ID, 'btnRemove')

kupot_kind_list = [(hishtalmut_clali, 'השתלמות כללי'), 
                   (hishtalmut_stocks, 'השתלמות מנייתי'),
                   (gemel_clali, 'גמל להשקעה כללי'),
                   (gemel_stocks, 'גמל להשקעה מנייתי'),
                   (hishtalmut_sp, 'השתלמות עוקב מדד S&P 500'),
                   (gemel_sp, 'גמל להשקעה עוקב מדד S&P 500')]
for group, subgroup in kupot_kind_list:
    yields_extractor(group, subgroup)

time.sleep(4) #just to see result
driver.quit()
   

hishtalmut_files = [('השתלמות כללי_yields.xls', 'השתלמות כללי_alpha.xls'),
                    ('השתלמות מנייתי_yields.xls', 'השתלמות מנייתי_alpha.xls'),
                    ('השתלמות עוקב מדד S&P 500_yields.xls', 'השתלמות עוקב מדד S&P 500_alpha.xls')]
gemel_files = [('גמל להשקעה כללי_yields.xls', 'גמל להשקעה כללי_alpha.xls'),
               ('גמל להשקעה מנייתי_yields.xls', 'גמל להשקעה מנייתי_alpha.xls'),
               ('גמל להשקעה עוקב מדד S&P 500_yields.xls', 'גמל להשקעה עוקב מדד S&P 500_alpha.xls')]

def process_files(files_list):
    dataframes = []
    for yields_file, alpha_file in files_list:
        gemel_yields = pd.read_excel(os.path.join(download_path, yields_file), skiprows=14, nrows=13)
        gemel_alpha = pd.read_excel(os.path.join(download_path, alpha_file), skiprows=12)
        alpha_data = gemel_alpha.loc[:, ['מספר', 'אלפא\nשנתי']].drop(index=0)
        alpha_data.rename(columns={'אלפא\nשנתי': 'אלפא שנתי'}, inplace=True)

        new_header = gemel_yields.columns.to_series().str.cat(gemel_yields.iloc[0], sep='_', na_rep='')
        gemel_yields.columns = new_header
        gemel_yields = gemel_yields.drop(gemel_yields.index[0])
        gemel_yields.reset_index(drop=True, inplace=True)
        gemel_yields = gemel_yields.drop(0)
        gemel_yields = gemel_yields.iloc[:, [1, 3, 4, 6, 7, 9, 10, 12, 14, 15, 18, 20]]

        # שמירת שורת הסיכום לפני עיבוד
        last_row = gemel_yields.iloc[-1:].copy()

        # עיבוד רק השורות עם נתונים (ללא שורת הסיכום)
        data_rows = gemel_yields[:-1].copy()

        # המרה למספרים והצטרפות לנתוני אלפא
        data_rows['מצטברת\nלתקופת\nהדו"ח_'] = pd.to_numeric(data_rows['מצטברת\nלתקופת\nהדו"ח_'], errors='coerce')

        # מיזוג עם נתוני אלפא לפני המיון
        data_rows = data_rows.merge(alpha_data, left_on='מספר_', right_on='מספר', how='left')
        data_rows = data_rows.drop(columns=['מספר'])

        # מיון לפי תשואה מצטברת
        sorted_df = data_rows.sort_values(by='מצטברת\nלתקופת\nהדו"ח_', ascending=False).copy()

        # הוספת שורת הסיכום בסוף
        sorted_df = pd.concat([sorted_df, last_row], ignore_index=False)
        sorted_df.fillna("", inplace=True)

        # שינוי שמות העמודות
        sorted_df.columns = sorted_df.columns.str.replace('מדד נזילות\n_', 'מדד נזילות')
        sorted_df.columns = sorted_df.columns.str.replace('צבירה\n נטו\n*_', 'צבירה נטו')
        sorted_df.columns = sorted_df.columns.str.replace('יתרת נכסים לסוף התקופה _', 'יתרת נכסים לסוף תקופה')
        sorted_df.columns = sorted_df.columns.str.replace('שיעור דמי ניהול ממוצע לשנת\n2024_נכסים', 'שיעור דמי ניהול ממוצע לשנת 2024 נכסים')
        sorted_df.columns = sorted_df.columns.str.replace('Unnamed: 7_הפקדות', 'שיעור דמי ניהול ממוצע לשנת 2024 הפקדות')
        sorted_df.columns = sorted_df.columns.str.replace('שארפ\nריבית\nחסרת \nסיכון_בנקודות האחוז', 'שארפ ריבית חסרת סיכון')
        sorted_df.columns = sorted_df.columns.str.replace('ממוצעת שנתית 5 שנים אחרונות_', 'ממוצעת שנתית 5 שנים אחרונות')
        sorted_df.columns = sorted_df.columns.str.replace('ממוצעת שנתית 3 שנים אחורונות_', 'ממוצעת שנתית 3 שנים אחרונות')
        sorted_df.columns = sorted_df.columns.str.replace('מצטברת\nלתקופת\nהדו"ח_', 'מצטברת לתקופת דיווח')
        sorted_df.columns = sorted_df.columns.str.replace('תקופת \nדיווח_', 'תקופת דיווח')
        sorted_df.columns = sorted_df.columns.str.replace('שם קופה_', 'שם קופה')
        sorted_df.columns = sorted_df.columns.str.replace('מספר_', 'מספר קופה')

        # סידור העמודות מחדש
        sorted_df = sorted_df.reindex(columns=['אלפא שנתי', 'מדד נזילות', 'צבירה נטו', 'יתרת נכסים לסוף תקופה',
                                               'שיעור דמי ניהול ממוצע לשנת 2024 נכסים',
                                               'שיעור דמי ניהול ממוצע לשנת 2024 הפקדות', 'שארפ ריבית חסרת סיכון',
                                               'ממוצעת שנתית 5 שנים אחרונות', 'ממוצעת שנתית 3 שנים אחרונות',
                                               'מצטברת לתקופת דיווח', 'תקופת דיווח', 'שם קופה', 'מספר קופה'])
        sorted_df['אלפא שנתי'].fillna('', inplace=True)
        sorted_df['שם קופה'] = sorted_df['שם קופה'].astype(str).str.replace(r'^\s*\*{3}\s*', '', regex=True)

        dataframes.append(sorted_df)

    return dataframes


# עיבוד הקבצים
hishtalmut_dataframes = process_files(hishtalmut_files)
gemel_dataframes = process_files(gemel_files)

def format_and_save_excel(dataframes, output_path):
    """שמירת שתי dataframes לקובץ אקסל אחד בגיליון אחד עם 2 שורות הפרדה"""
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        workbook = writer.book
        sheet = workbook.create_sheet('תשואות')

        # עיצוב כותרות
        header_fill = PatternFill(start_color='001054', end_color='001054', fill_type='solid')
        header_font = Font(name='david', bold=True, color='FFFFFF')
        text_font = Font(name='david')
        sum_row_fill = PatternFill(start_color='4859fb', end_color='4859fb', fill_type='solid')
        sum_row_font = Font(name='david', bold=True, color='FFFFFF')

        current_row = 1

        for df_idx, df in enumerate(dataframes):
            # כתיבת כותרות
            for col_idx, column_name in enumerate(df.columns, start=1):
                cell = sheet.cell(row=current_row, column=col_idx, value=column_name)
                cell.fill = header_fill
                cell.font = header_font

            # כתיבת נתונים
            for row_idx, row_data in enumerate(df.values, start=current_row + 1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = text_font

                    # יישור למרכז
                    column_letter = get_column_letter(col_idx)
                    column_to_align = ['A', 'B','C','D','E','F','G','H','I', 'J', 'L', 'M']
                    if column_letter in column_to_align:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # הוספת פסיקים למספרים
                    column_to_add_commas = ['B', 'C']
                    if column_letter in column_to_add_commas and isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

            # עיצוב שורת סיכום (השורה האחרונה של כל טבלה)
            last_data_row = current_row + len(df)
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=last_data_row, column=col_idx)
                cell.fill = sum_row_fill
                cell.font = sum_row_font

            # מעבר לטבלה הבאה (שורת כותרת + נתונים + 2 שורות הפרדה)
            current_row = last_data_row + 3  # +1 לסיום הטבלה הנוכחית + 2 שורות הפרדה

        # מחיקת גיליון ברירת המחדל אם קיים
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

# שמירת קבצי אקסל
format_and_save_excel(
    hishtalmut_dataframes,
    os.path.join(download_path, 'השתלמות_תשואות_מעובד.xlsx')
)

format_and_save_excel(
    gemel_dataframes,
    os.path.join(download_path, 'גמל_תשואות_מעובד.xlsx')
)

print("הקבצים נשמרו בהצלחה!")
print(f"השתלמות: {os.path.join(download_path, 'השתלמות_תשואות_מעובד.xlsx')}")
print(f"גמל: {os.path.join(download_path, 'גמל_תשואות_מעובד.xlsx')}")
